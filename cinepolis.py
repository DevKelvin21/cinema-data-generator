from __future__ import annotations

import argparse
import importlib
import pkgutil
from datetime import date
from typing import Iterable, List, Tuple
import pandas as pd
from openpyxl.utils import get_column_letter

PACKAGE_NAME = "cinepolis_ca"
MODULE_PREFIX = "cinepolis_"
REQUIRED_COLUMNS = ["Country", "Theater", "Date", "Time", "Movie", "Format"]

SPANISH_HEADER_MAP = {
    "Country": "Pais",
    "Theater": "Nombre del cine",
    "Date": "Fecha de funcion",
    "Time": "hora de funcion",
    "Movie": "Nombre de la pelicula",
    "Format": "Formato de la pelicula",
}

ENGLISH_HEADER_MAP = {
    "Country": "Country",
    "Theater": "Theater",
    "Date": "Date",
    "Time": "Time",
    "Movie": "Movie",
    "Format": "Format",
}

def _find_country_modules() -> List[str]:
    """Lista nombres de módulos cinepolis_*.py en cinepolis_ca (sin el .py)."""
    package = importlib.import_module(PACKAGE_NAME)
    modules = []
    for m in pkgutil.iter_modules(package.__path__):
        if m.name.startswith(MODULE_PREFIX):
            modules.append(m.name)  # e.g. "cinepolis_gt"
    return sorted(modules)

def _import_scraper(module_name: str):
    """Importa cinepolis_ca.<module_name> y retorna su objeto módulo."""
    full = f"{PACKAGE_NAME}.{module_name}"
    return importlib.import_module(full)

def _parse_date_column(s: pd.Series) -> pd.Series:
    """Acepta YYYY-MM-DD, YYYY/MM/DD, DD/MM/YYYY; devuelve objetos date."""
    parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
    return parsed.dt.date

def _normalize(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}. Got: {list(df.columns)}")

    out = df[REQUIRED_COLUMNS].copy()
    for c in ["Country", "Theater", "Time", "Movie", "Format"]:
        out[c] = out[c].astype(str).str.strip()
    out["Date"] = _parse_date_column(out["Date"])
    out = out.dropna(subset=["Date"])
    out["Date"] = out["Date"].astype(object)  # python date objects (Excel-friendly)
    return out

def _split_today_other(df: pd.DataFrame, next_days: int = 7) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return df.copy(), df.copy()
    distinct = sorted({d for d in df["Date"] if isinstance(d, date)})
    if not distinct:
        return df.copy(), pd.DataFrame(columns=df.columns)
    today = distinct[0]
    nexts = distinct[1:1 + next_days]
    return df[df["Date"] == today].copy(), df[df["Date"].isin(nexts)].copy()

def _apply_excel_date_format(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, date_col: str, fmt: str = "DD/MM/YYYY"):
    """Aplica formato de fecha a la columna indicada en una hoja específica."""
    if sheet_name not in writer.sheets or df.empty or date_col not in df.columns:
        return
    ws = writer.sheets[sheet_name]
    col_idx = df.columns.get_loc(date_col) + 1
    col_letter = get_column_letter(col_idx)
    for row in range(2, len(df) + 2):  # data rows (header = row 1)
        ws[f"{col_letter}{row}"].number_format = fmt

def _rename_headers(df: pd.DataFrame, lang: str) -> pd.DataFrame:
    """Renombra encabezados según idioma (es/en)."""
    header_map = SPANISH_HEADER_MAP if lang == "es" else ENGLISH_HEADER_MAP
    return df.rename(columns=header_map)

def combine_and_save(
    frames: Iterable[pd.DataFrame],
    out_path: str = "cinepolis_combined.xlsx",
    next_days: int = 7,
    date_fmt: str = "DD/MM/YYYY",
    keep_all: bool = False,
    lang: str = "es",
) -> str:
    normalized = []
    for i, df in enumerate(frames, start=1):
        if df is None or df.empty:
            continue
        normalized.append(_normalize(df))

    # Si no hay datos, crear archivo vacío con headers correctos y hojas en español
    if not normalized:
        empty = pd.DataFrame(columns=REQUIRED_COLUMNS)
        empty_es = _rename_headers(empty, lang=lang)
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            sheet_today = "Hoy" if lang == "es" else "Today"
            sheet_other = "Otras fechas" if lang == "es" else "Other schedules"
            empty_es.to_excel(w, sheet_name=sheet_today, index=False)
            empty_es.to_excel(w, sheet_name=sheet_other, index=False)
        return out_path

    all_df = pd.concat(normalized, ignore_index=True)

    # Split por fechas (hoy y siguientes N fechas)
    today_df, other_df = _split_today_other(all_df, next_days=next_days)

    # Renombrar encabezados al idioma elegido ANTES de escribir
    today_es = _rename_headers(today_df, lang=lang)
    other_es = _rename_headers(other_df, lang=lang)
    all_es = _rename_headers(all_df, lang=lang) if keep_all else None

    # Nombres de hojas en español por defecto
    sheet_today = "Hoy" if lang == "es" else "Today"
    sheet_other = "Otras fechas" if lang == "es" else "Other schedules"
    sheet_all = "Todo" if lang == "es" else "All"

    # Calcular nombre de la columna de fecha ya renombrada
    date_col_name = SPANISH_HEADER_MAP["Date"] if lang == "es" else ENGLISH_HEADER_MAP["Date"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        today_es.to_excel(w, sheet_name=sheet_today, index=False)
        _apply_excel_date_format(w, sheet_today, today_es, date_col_name, date_fmt)

        other_es.to_excel(w, sheet_name=sheet_other, index=False)
        _apply_excel_date_format(w, sheet_other, other_es, date_col_name, date_fmt)

        if keep_all and all_es is not None:
            all_es.to_excel(w, sheet_name=sheet_all, index=False)
            _apply_excel_date_format(w, sheet_all, all_es, date_col_name, date_fmt)

    return out_path

def main():
    parser = argparse.ArgumentParser(description="Combina scrapers de Cinépolis en un solo Excel.")
    parser.add_argument("--countries", type=str, default="all",
                        help="Sufijos separados por coma (ej. 'gt,el_salvador') o 'all'")
    parser.add_argument("--output", type=str, default="cinepolis_combined.xlsx")
    parser.add_argument("--headless", action="store_true", default=True)
    parser.add_argument("--no-headless", dest="headless", action="store_false")
    parser.add_argument("--keep-all", action="store_true", help="Agrega hoja 'Todo' con todas las filas.")
    parser.add_argument("--next-days", type=int, default=7)
    parser.add_argument("--lang", type=str, choices=["es", "en"], default="es", help="Idioma de encabezados y hojas.")
    args = parser.parse_args()

    # Descubrir módulos disponibles
    available = _find_country_modules()            # ['cinepolis_gt', 'cinepolis_el_salvador', ...]
    wanted = available

    if args.countries.lower() != "all":
        suffixes = {s.strip() for s in args.countries.split(",") if s.strip()}
        wanted = [m for m in available if m.replace(MODULE_PREFIX, "") in suffixes]

        unknown = suffixes - {m.replace(MODULE_PREFIX, "") for m in available}
        if unknown:
            print("Warning: unknown modules:", ", ".join(sorted(unknown)))

    frames: List[pd.DataFrame] = []
    for mod_name in wanted:
        try:
            mod = _import_scraper(mod_name)
            if not hasattr(mod, "scrape"):
                print(f"Skipping {mod_name}: no scrape()")
                continue
            print(f"Running {mod_name}.scrape(headless={args.headless}) ...")
            df = mod.scrape(headless=args.headless)
            if df is not None and not df.empty:
                frames.append(df)
                print(f" -> {mod_name}: {len(df)} rows")
            else:
                print(f" -> {mod_name}: empty")
        except Exception as e:
            print(f"Error in {mod_name}: {e}")

    out = combine_and_save(
        frames,
        out_path=args.output,
        next_days=args.next_days,
        keep_all=args.keep_all,
        lang=args.lang,
        date_fmt="DD/MM/YYYY",  # fuerza dd/mm/yyyy en Excel
    )
    print("Guardado:", out)

if __name__ == "__main__":
    main()
